# -*- coding: utf-8 -*-
"""
Excel导出模块
将发票识别结果导出到Excel文件
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime

from invoice_parser import InvoiceInfo


class ExcelExporter:
    """Excel导出器类"""
    
    # 表头定义
    HEADERS = [
        ('序号', 8),
        ('文件名', 25),
        ('发票代码', 15),
        ('发票号码', 18),
        ('金额(元)', 12),
        ('税率', 8),
        ('税额(元)', 12),
        ('价税合计(元)', 14),
        ('开票日期', 12),
        ('销售方', 30),
        ('备注', 20)
    ]
    
    def __init__(self):
        """初始化导出器"""
        self.workbook = None
        self.worksheet = None
    
    def export(self, invoices: List[InvoiceInfo], output_path: str) -> bool:
        """
        导出发票信息到Excel
        
        Args:
            invoices: 发票信息列表
            output_path: 输出文件路径
            
        Returns:
            是否导出成功
        """
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "发票汇总"
            
            # 设置表头
            self._set_headers()
            
            # 写入数据
            for idx, invoice in enumerate(invoices, start=1):
                self._write_invoice_row(idx, invoice)
            
            # 写入汇总行
            self._write_summary_row(len(invoices))
            
            # 调整列宽
            self._adjust_column_widths()
            
            # 保存文件
            self.workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"导出Excel错误: {e}")
            return False
        
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _set_headers(self):
        """设置表头"""
        # 标题行样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, (header_name, width) in enumerate(self.HEADERS, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # 冻结首行
        self.worksheet.freeze_panes = 'A2'
    
    def _write_invoice_row(self, row_num: int, invoice: InvoiceInfo):
        """写入单行发票数据"""
        row = row_num + 1  # 跳过表头
        
        data = [
            row_num,                    # 序号
            invoice.file_name,          # 文件名
            invoice.invoice_code,       # 发票代码
            invoice.invoice_number,     # 发票号码
            invoice.amount,             # 金额
            invoice.tax_rate,           # 税率
            invoice.tax_amount,         # 税额
            invoice.total_amount,       # 价税合计
            invoice.invoice_date,       # 开票日期
            invoice.seller_name,        # 销售方
            invoice.remarks             # 备注
        ]
        
        # 样式
        alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, value in enumerate(data, start=1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            # 根据列类型设置对齐方式
            if col in [5, 7, 8]:  # 金额列右对齐
                cell.alignment = right_alignment
                cell.number_format = '#,##0.00'
            elif col in [2, 10]:  # 文件名和销售方左对齐
                cell.alignment = left_alignment
            else:
                cell.alignment = alignment
    
    def _write_summary_row(self, invoice_count: int):
        """写入汇总行"""
        if invoice_count == 0:
            return
        
        row = invoice_count + 2  # 表头 + 数据行数
        
        # 合计标签
        cell = self.worksheet.cell(row=row, column=1, value="合计")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并前四个单元格
        self.worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # 金额汇总公式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 金额合计
        cell = self.worksheet.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 税率列空白
        cell = self.worksheet.cell(row=row, column=6, value="")
        cell.border = thin_border
        
        # 税额合计
        cell = self.worksheet.cell(row=row, column=7, value=f"=SUM(G2:G{row-1})")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 价税合计
        cell = self.worksheet.cell(row=row, column=8, value=f"=SUM(H2:H{row-1})")
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # 其他空白单元格也加边框
        for col in [1, 2, 3, 4, 9, 10, 11]:
            cell = self.worksheet.cell(row=row, column=col)
            cell.border = thin_border
    
    def _adjust_column_widths(self):
        """调整列宽（已在设置表头时完成）"""
        pass


def export_to_excel(invoices: List[InvoiceInfo], output_path: str) -> bool:
    """
    便捷函数：导出发票到Excel
    
    Args:
        invoices: 发票信息列表
        output_path: 输出文件路径
        
    Returns:
        是否成功
    """
    exporter = ExcelExporter()
    return exporter.export(invoices, output_path)


def generate_output_filename(directory: str = None) -> str:
    """
    生成输出文件名
    
    Args:
        directory: 输出目录（可选）
        
    Returns:
        输出文件完整路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"发票汇总_{timestamp}.xlsx"
    
    if directory:
        from pathlib import Path
        return str(Path(directory) / filename)
    
    return filename
