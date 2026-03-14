"""
Excel导出模块

该模块负责将解析后的ARXML数据导出为格式化的Excel文件。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from pathlib import Path


class ExcelExporter:
    """Excel导出器"""
    
    # 样式定义
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    HEADER_FILL = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    CELL_ALIGNMENT = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 交替行颜色
    ALT_ROW_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    def __init__(self):
        """初始化导出器"""
        self.workbook = Workbook()
        
    def export(self, data: Dict[str, Any], output_path: str) -> bool:
        """
        导出数据到Excel文件
        
        Args:
            data: 解析后的ARXML数据
            output_path: 输出文件路径
            
        Returns:
            bool: 导出成功返回True
        """
        try:
            # 删除默认工作表
            if 'Sheet' in self.workbook.sheetnames:
                del self.workbook['Sheet']
            
            # 创建概览表
            self._create_overview_sheet(data)
            
            # 按元素类型创建详细表
            self._create_element_sheets(data)
            
            # 创建包结构表
            self._create_packages_sheet(data)
            
            # 保存文件
            self.workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"导出失败: {e}")
            return False
    
    def _create_overview_sheet(self, data: Dict[str, Any]):
        """创建概览工作表"""
        ws = self.workbook.create_sheet('概览', 0)
        
        # 统计信息
        stats = self._get_statistics(data)
        
        # 写入标题
        ws['A1'] = 'ARXML文件概览'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # 写入统计信息
        row = 3
        headers = ['统计项', '数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        row += 1
        for stat_name, count in stats.items():
            ws.cell(row=row, column=1, value=stat_name).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=count).border = self.THIN_BORDER
            if (row - 4) % 2 == 0:
                ws.cell(row=row, column=1).fill = self.ALT_ROW_FILL
                ws.cell(row=row, column=2).fill = self.ALT_ROW_FILL
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _get_statistics(self, data: Dict[str, Any]) -> Dict[str, int]:
        """获取统计信息"""
        stats = {
            '总包数': 0,
            '总元素数': 0,
        }
        
        element_types = {}
        
        def count_package(pkg: Dict[str, Any]):
            stats['总包数'] += 1
            
            for elem in pkg.get('elements', []):
                stats['总元素数'] += 1
                elem_type = elem.get('type', 'Unknown')
                element_types[elem_type] = element_types.get(elem_type, 0) + 1
            
            for sub_pkg in pkg.get('sub_packages', []):
                count_package(sub_pkg)
        
        for pkg in data.get('packages', []):
            count_package(pkg)
        
        # 添加元素类型统计
        for elem_type, count in sorted(element_types.items()):
            stats[f'  - {elem_type}'] = count
        
        return stats
    
    def _create_element_sheets(self, data: Dict[str, Any]):
        """按元素类型创建工作表"""
        # 收集所有元素
        all_elements = []
        
        def collect_elements(pkg: Dict[str, Any], path: str = ''):
            current_path = f"{path}/{pkg['short_name']}" if path else pkg['short_name']
            
            for elem in pkg.get('elements', []):
                elem['_package_path'] = current_path
                all_elements.append(elem)
            
            for sub_pkg in pkg.get('sub_packages', []):
                collect_elements(sub_pkg, current_path)
        
        for pkg in data.get('packages', []):
            collect_elements(pkg)
        
        if not all_elements:
            return
        
        # 按类型分组
        elements_by_type = {}
        for elem in all_elements:
            elem_type = elem.get('type', 'Other')
            if elem_type not in elements_by_type:
                elements_by_type[elem_type] = []
            elements_by_type[elem_type].append(elem)
        
        # 为每种类型创建工作表
        for elem_type, elements in elements_by_type.items():
            # 限制工作表名称长度（Excel限制31字符）
            sheet_name = elem_type[:28] if len(elem_type) > 28 else elem_type
            # 处理无效字符
            sheet_name = sheet_name.replace('/', '-').replace('\\', '-')
            sheet_name = sheet_name.replace('[', '(').replace(']', ')')
            sheet_name = sheet_name.replace(':', '-').replace('*', '-')
            sheet_name = sheet_name.replace('?', '-').replace("'", '-')
            
            ws = self.workbook.create_sheet(sheet_name)
            self._write_elements_to_sheet(ws, elements)
    
    def _write_elements_to_sheet(self, ws, elements: List[Dict[str, Any]]):
        """将元素写入工作表"""
        if not elements:
            return
        
        # 定义列
        columns = [
            ('包路径', '_package_path'),
            ('短名称', 'short_name'),
            ('长名称', 'long_name'),
            ('描述', 'description'),
            ('类别', 'category'),
        ]
        
        # 收集所有属性键
        all_attr_keys = set()
        for elem in elements:
            attrs = elem.get('attributes', {})
            for key, value in attrs.items():
                if isinstance(value, (str, int, float)):
                    all_attr_keys.add(key)
        
        # 添加属性列
        for key in sorted(all_attr_keys):
            columns.append((key, f'_attr_{key}'))
        
        # 写入表头
        for col, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 写入数据
        for row_idx, elem in enumerate(elements, 2):
            # 预处理属性
            attrs = elem.get('attributes', {})
            for key, value in attrs.items():
                if isinstance(value, (str, int, float)):
                    elem[f'_attr_{key}'] = value
            
            for col, (_, key) in enumerate(columns, 1):
                value = elem.get(key, '')
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = str(value)
                
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = self.CELL_ALIGNMENT
                cell.border = self.THIN_BORDER
                
                # 交替行颜色
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
        
        # 自动调整列宽
        for col in range(1, len(columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, len(elements) + 2):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, min(cell_length, 50))
            
            ws.column_dimensions[column_letter].width = max(max_length + 2, 10)
    
    def _create_packages_sheet(self, data: Dict[str, Any]):
        """创建包结构工作表"""
        ws = self.workbook.create_sheet('包结构')
        
        # 表头
        headers = ['包路径', '包名称', '长名称', '描述', '元素数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 收集包信息
        row = 2
        
        def write_package(pkg: Dict[str, Any], path: str = '', level: int = 0):
            nonlocal row
            current_path = f"{path}/{pkg['short_name']}" if path else pkg['short_name']
            
            # 缩进显示层级
            indent = '  ' * level
            
            ws.cell(row=row, column=1, value=current_path).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=f"{indent}{pkg['short_name']}").border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=pkg.get('long_name', '')).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=pkg.get('description', '')).border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=len(pkg.get('elements', []))).border = self.THIN_BORDER
            
            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = self.ALT_ROW_FILL
            
            row += 1
            
            for sub_pkg in pkg.get('sub_packages', []):
                write_package(sub_pkg, current_path, level + 1)
        
        for pkg in data.get('packages', []):
            write_package(pkg)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12


def export_to_excel(data: Dict[str, Any], output_path: str) -> bool:
    """
    导出到Excel的便捷函数
    
    Args:
        data: 解析后的ARXML数据
        output_path: 输出文件路径
        
    Returns:
        bool: 导出成功返回True
    """
    exporter = ExcelExporter()
    return exporter.export(data, output_path)
